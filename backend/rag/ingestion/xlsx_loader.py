import openpyxl
from typing import List, Dict, Any
from rag.ingestion.loader import BaseLoader
from rag.schemas.document import PageContent

class XLSXLoader(BaseLoader):
    def load(self, file_path: str) -> List[PageContent]:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        pages = []
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            rows_data = []
            for row in ws.iter_rows(values_only=True):
                if any(row):
                    rows_data.append(" | ".join([str(c) if c is not None else "" for c in row]))
            
            sheet_text = f"Sheet: {sheet_name}\n" + "\n".join(rows_data)
            pages.append(PageContent(
                page_number=sheet_idx + 1,
                text=sheet_text,
                tables=[{"sheet_name": sheet_name, "row_count": len(rows_data)}],
                metadata={"sheet_name": sheet_name}
            ))
        wb.close()
        return pages
